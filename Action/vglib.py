import fetchdb, report_hard_code_data
import time, csv, datetime, argparse, pickle, os
from calendar import monthrange
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

def singleton(C):
	def wrapper(*args, **kargs):
		if not hasattr(C, 'instance'):
			C.instance = C(*args, **kargs)
		return C.instance
	return wrapper


vglib_params = {
	'gs_app_key_file_name': 'E:/Home/mh/vg_new/config/app_key.json',
	'app_auto_report_secret': '2681bdecbc7aee70f0e6f017705d1e20',	
}


def gspread_auth(json_key=vglib_params['gs_app_key_file_name'], scope=['https://spreadsheets.google.com/feeds']):
	credentials = ServiceAccountCredentials.from_json_keyfile_name(json_key, scope)
	gc = gspread.authorize(credentials)
	return gc

# default fetch the config sheet
def get_gs_sheet(file_id='', sheet_name='', gs_app_key=vglib_params['gs_app_key_file_name']):
	gc = gspread_auth(gs_app_key) 
	wsh = gc.open_by_key(file_id)
	sh = wsh.worksheet(sheet_name)	
	return sh	
	
# find which orders_product an estore is associated
def find_op_table(estore_id):
	t = int(estore_id) % 20
	return 'orders_product_{}'.format(t)

# return dict containing records being placed in the correct orders_product tables	
# index is the index of what is stored in the dict. Most of the time, it is ord_id
def get_op_tables(records, estore_index, index):
	op_tables = {
		'orders_product_0':[],
		'orders_product_1':[],
		'orders_product_2':[],
		'orders_product_3':[],
		'orders_product_4':[],
		'orders_product_5':[],
		'orders_product_6':[],
		'orders_product_7':[],
		'orders_product_8':[],
		'orders_product_9':[],
		'orders_product_10':[],
		'orders_product_11':[],
		'orders_product_12':[],
		'orders_product_13':[],
		'orders_product_14':[],
		'orders_product_15':[],
		'orders_product_16':[],
		'orders_product_17':[],
		'orders_product_18':[],
		'orders_product_19':[],
	}
	
	# determine orders_product_x 
	# append only the r[index] (supposed to be the ord_id) into corresponding list
	for r in records:
		op_tables[find_op_table(r[estore_index])].append(r[index])
		
	return op_tables

# atime is like '2016-01'	
def convert_to_datetime(atime):
	y,m = atime.split('-')

	first = '{}-{}-01 00:00:00'.format(y,m)
	last = '{}-{}-{} 23:59:59'.format(y,m, monthrange(int(y), int(m))[1])
	return first,last			
		
# fetch records using ord_id.
# return records, NOT write to a file 
def fetch_order_with_ord_id(query, year_month='', heavy=1):
	if year_month:
		start, end = convert_to_datetime(year_month)
		heavy_start,heavy_end = fetchdb.gotHeavyParam('ord_id', 'orders_new', 'ord_date', start, end)
	if heavy:
		records = fetchdb.select_mh(query=query, is_csv=0, is_json=0, fname='', heavy=1, heavy_attr='orders_new.ord_id', heavy_start=heavy_start, heavy_end=heavy_end, mesg='', title='')
	else:
		records = fetchdb.select_mh(query=query, is_csv=0, is_json=0, fname='', heavy=heavy)
	return records	


# assume year_month_list is sorted in ascending order like ['2016-01', '2016-02', '2016-03']	
def fetch_order_with_ord_id_all(query, year_month_list='', heavy=1):
	num = len(year_month_list)
	start1, end1 = convert_to_datetime(year_month_list[0])
	heavy_start1,heavy_end1 = fetchdb.gotHeavyParam('ord_id', 'orders_new', 'ord_date', start1, end1)
	
	
	start2, end2 = convert_to_datetime(year_month_list[num-1])
	heavy_start2,heavy_end2 = fetchdb.gotHeavyParam('ord_id', 'orders_new', 'ord_date', start2, end2)	
	
	if heavy:
		records = fetchdb.select_mh(query=query, is_csv=0, is_json=0, fname='', heavy=heavy, heavy_attr='orders_new.ord_id', heavy_start=heavy_start1, heavy_end=heavy_end2, mesg='', title='')
	else:
		records = fetchdb.select_mh(query=query, is_csv=0, is_json=0, fname='', heavy=heavy)		
	return records	
	

# fetch and join results from orders_product, product_multi, and other tables   
# in the query, the first value is the number of orders_product, and the second is the list of order_id	
# NOT USE HEAVY MODE in fetching 
def fetch_order_product(op_tables, query, is_csv=0, short_name='', other_func='', fname=''):
	if short_name:
		fname = gen_fname(short_name)
		
	record_total = []
	for t in op_tables:
		if op_tables[t]:
			id_list = ','.join(map(str, op_tables[t]))
		
			tempquery = query.format(t, id_list)	

			try:
				records = fetchdb.select_mh(query=tempquery, is_csv=0, is_json=0, fname='', heavy=0, heavy_attr='')
				
			except Exception as ex:
				print(ex)
				print(tempquery)
				exit()
			else:
				if not is_csv: 
					record_total = record_total + list(records)
				else:
					if other_func:
						record_total = record_total + list(records)
					else:
						to_csv(records, fname, 'a')
	if other_func:
		record_fixed = other_func(record_total)
		to_csv(record_fixed, fname, 'a')
	
	return record_total	if not is_csv else fname	

# generate unique file name 
def gen_fname(short_name, prefix='.output'):
	return report_hard_code_data.path + str(int(time.time()))+'_'+short_name+prefix
	
def create_dir_if_not_exist(dir='outputs'):
	if not os.path.isdir(dir):
		os.makedirs(dir)
		
	
def to_csv(records, fname, mode='w'):
	create_dir_if_not_exist(fname.split('/')[0]) # assuming fname is in format 'dir/filename'

	
	with open(fname, mode, newline='', encoding='utf-8') as fd:
		writer = csv.writer(fd, delimiter=' ')	
		for r in records:
			writer.writerow(r)

# auto generate file name, given the short name of the file			
def to_csv_2(records, shortName, mode='w'):
	fname = gen_fname(shortName)
	to_csv(records, fname, mode)
	return fname

# assume path exits
# fetch all row in a csv file and return result  	
def from_csv(fname, mode='r'):
	result = []
	
	with open(fname, mode, newline='', encoding='utf-8') as fd:
		reader = csv.reader(fd, delimiter=' ')	
		for r in reader:
			r = list(r)
			result.append(r)
				
	return result
	
def from_csv_gen(fname, mode='r'):
	with open(fname, mode, newline='', encoding='utf-8') as fd:
		reader = csv.reader(fd, delimiter=' ')
		for record in reader:
			yield record
	
# mode determines the return is writer or reader			
def get_csv_object(fname, mode='w'):
	create_dir_if_not_exist(fname.split('/')[0])
	fd = open(fname, mode, newline='', encoding='utf-8')
	if mode == 'w' or mode == 'a':
		csver = csv.writer(fd, delimiter=' ')	
	elif mode == 'r':
		csver = csv.reader(fd, delimiter=' ')
	return [fd, csver]		
			
def to_excel(records, file_name, sheet_name, new=1, title=''):
	if new:
		wb = openpyxl.Workbook()
	else:
		wb = openpyxl.load_workbook(file_name)
	sh = wb.create_sheet(index=0, title=sheet_name)
	
	recordNum = len(records)
	
	if title:
		num1 = 1
		sh.cell(row=1, column=1).value = title
	else:
		num1 = 0
	
	for i, row in zip(range(1+num1, recordNum+1+num1),records):
		rlen = len(row)
		for j in range(1, rlen+1):
			sh.cell(row=i, column=j).value = row[j-1]
	
	wb.save(file_name)			
			
# mapping main and sub categories
# records is a list of tuples of data
# return a dict, whose key are main cat, and whose values are a dict with keys being estore ids and values being (amount, order num) 
# keyIndex, in most case, is the index of the estore id in each record			
def mapping_main_sub_cat(records, keyIndex, aggregate_func, int_insert_func, subIndex, main_sub_mapping_file):
	main_cat = {
		1:{},
		2:{},
		3:{},
		4:{},
		5:{},
		6:{},
		7:{},
		8:{},
		9:{},
		10:{},
		11:{},
		12:{},
		13:{},
		14:{},
		15:{},
		16:{},
		17:{},
		18:{},
		19:{},
		20:{},
		21:{},
		22:{}
	}	
	
	with open(main_sub_mapping_file, 'rb') as fd:
		d = pickle.load(fd)
		
		for row in records:
			subId = int(row[subIndex])
			main_id = d[subId]
			estore_id = int(row[keyIndex])
			
			if main_cat[main_id].get(estore_id, -1) == -1:
				# insert value if the keyIndex not exists in the dict	
				main_cat[main_id][estore_id] = []
				int_insert_func(row,main_cat[main_id]) 

			else:
				# aggregate value if the keyIndex has been existed 
				aggregate_func(row,main_cat[main_id])
		
	return main_cat				
			
			
# mapping main and sub categories 
def build_main_sub_category_file():
	wb = openpyxl.load_workbook('mapping_cat_sub_main.xlsx', data_only=True)
	sh = wb.get_sheet_by_name('Sheet1')
	
	maxrow = sh.max_row
	
	d = {}
	
	with open('main_sub_category_mapping.csv', 'w', newline='', encoding='utf-8') as fd:
		writer = csv.writer(fd, delimiter=' ')
		
		for i in range(3,maxrow + 1):
			if int(sh.cell(row=i,column=6).value) != -1:
				print('found row',i)
				sub = int(sh.cell(row=i,column=1).value)
				main = int(sh.cell(row=i,column=6).value)
				writer.writerow((sub, main))
				d[sub] = main
	# a dict, whose key is the sub category id, and value is the main category id
	with open('main_sub_category_mapping.pkl', 'wb') as f:
		pickle.dump(d, f)

# add more mapping main and sub categories for pickle and csv
def add_main_sub_mapping(sub, main):
	with open('main_sub_category_mapping.csv', 'w', newline='', encoding='utf-8') as fd:
		writer = csv.writer(fd, delimiter=' ')	
		writer.writerow((sub, main))
	

	with open('main_sub_category_mapping.pkl', 'wb') as f:
		d = {sub:main}
		pickle.dump(d, f)	
		